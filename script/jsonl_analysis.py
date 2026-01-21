import json
import statistics
from collections import defaultdict
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def calculate_statistics(values: List[float]) -> Dict:
    """수치 리스트의 통계값을 계산"""
    if not values:
        return None
    
    sorted_values = sorted(values)
    n = len(values)
    
    stats = {
        'count': n,
        'min': min(values),
        'max': max(values),
        'mean': statistics.mean(values),
        'median': statistics.median(values),
        'stdev': statistics.stdev(values) if n > 1 else 0,
        'variance': statistics.variance(values) if n > 1 else 0,
        'q1': sorted_values[n // 4],
        'q3': sorted_values[3 * n // 4],
    }
    
    return stats

def extract_numeric_data(file_path: str, selected_fields: Optional[List[str]] = None) -> Dict[str, List[float]]:
    """JSONL 파일에서 수치 데이터를 추출"""
    numeric_data = defaultdict(list)
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                data = json.loads(line)
                
                for key, value in data.items():
                    # 선택된 필드가 있으면 해당 필드만, 없으면 모든 필드 처리
                    if selected_fields and key not in selected_fields:
                        continue
                    
                    if isinstance(value, (int, float)):
                        numeric_data[key].append(value)
    
    return numeric_data

def save_to_txt(numeric_data: Dict[str, List[float]], output_path: str, selected_fields: Optional[List[str]] = None):
    """분석 결과를 TXT 파일로 저장"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("JSONL 수치 데이터 분석 결과\n")
        f.write(f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        
        if selected_fields:
            f.write(f"선택된 필드 분석: {', '.join(selected_fields)}\n\n")
        else:
            f.write(f"전체 필드 분석 (총 {len(numeric_data)}개 필드)\n\n")
        
        for field_name, values in sorted(numeric_data.items()):
            stats = calculate_statistics(values)
            
            f.write("=" * 80 + "\n")
            f.write(f"필드: {field_name}\n")
            f.write("=" * 80 + "\n")
            f.write(f"  데이터 수:    {stats['count']}\n")
            f.write(f"  최솟값:       {stats['min']:.4f}\n")
            f.write(f"  최댓값:       {stats['max']:.4f}\n")
            f.write(f"  평균:         {stats['mean']:.4f}\n")
            f.write(f"  중앙값:       {stats['median']:.4f}\n")
            f.write(f"  1사분위수:    {stats['q1']:.4f}\n")
            f.write(f"  3사분위수:    {stats['q3']:.4f}\n")
            f.write(f"  표준편차:     {stats['stdev']:.4f}\n")
            f.write(f"  분산:         {stats['variance']:.4f}\n")
            f.write("\n")
    
    print(f"✓ TXT 파일 저장 완료: {output_path}")

def save_to_xlsx(numeric_data: Dict[str, List[float]], output_path: str, selected_fields: Optional[List[str]] = None):
    """분석 결과를 XLSX 파일로 저장"""
    wb = Workbook()
    ws = wb.active
    ws.title = "통계 분석 결과"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["필드명", "데이터 수", "최솟값", "최댓값", "평균", "중앙값", 
               "1사분위수", "3사분위수", "표준편차", "분산"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    row = 2
    for field_name, values in sorted(numeric_data.items()):
        stats = calculate_statistics(values)
        
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=2, value=stats['count'])
        ws.cell(row=row, column=3, value=round(stats['min'], 4))
        ws.cell(row=row, column=4, value=round(stats['max'], 4))
        ws.cell(row=row, column=5, value=round(stats['mean'], 4))
        ws.cell(row=row, column=6, value=round(stats['median'], 4))
        ws.cell(row=row, column=7, value=round(stats['q1'], 4))
        ws.cell(row=row, column=8, value=round(stats['q3'], 4))
        ws.cell(row=row, column=9, value=round(stats['stdev'], 4))
        ws.cell(row=row, column=10, value=round(stats['variance'], 4))
        
        row += 1
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    wb.save(output_path)
    print(f"✓ XLSX 파일 저장 완료: {output_path}")

def analyze_jsonl(file_path: str, selected_fields: Optional[List[str]] = None, 
                  save_txt: bool = True, save_xlsx: bool = True):
    """JSONL 파일 분석 및 통계 출력"""
    print(f"\n파일 분석 중: {file_path}\n")
    
    # 수치 데이터 추출
    numeric_data = extract_numeric_data(file_path, selected_fields)
    
    if not numeric_data:
        print("수치 데이터를 찾을 수 없습니다.")
        return
    
    # 분석 모드 출력
    if selected_fields:
        print(f"선택된 필드 분석: {', '.join(selected_fields)}\n")
    else:
        print(f"전체 필드 분석 (총 {len(numeric_data)}개 필드)\n")
    
    # 콘솔 출력
    for field_name, values in sorted(numeric_data.items()):
        stats = calculate_statistics(values)
        
        print("=" * 80)
        print(f"필드: {field_name}")
        print("=" * 80)
        print(f"  데이터 수:    {stats['count']}")
        print(f"  최솟값:       {stats['min']:.4f}")
        print(f"  최댓값:       {stats['max']:.4f}")
        print(f"  평균:         {stats['mean']:.4f}")
        print(f"  중앙값:       {stats['median']:.4f}")
        print(f"  1사분위수:    {stats['q1']:.4f}")
        print(f"  3사분위수:    {stats['q3']:.4f}")
        print(f"  표준편차:     {stats['stdev']:.4f}")
        print(f"  분산:         {stats['variance']:.4f}")
        print()
    
    # 파일 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = file_path.rsplit('.', 1)[0]
    
    print("\n" + "=" * 80)
    print("파일 저장 중...")
    print("=" * 80)
    
    if save_txt:
        txt_path = f"{base_name}_analysis_{timestamp}.txt"
        save_to_txt(numeric_data, txt_path, selected_fields)
    
    if save_xlsx:
        xlsx_path = f"{base_name}_analysis_{timestamp}.xlsx"
        save_to_xlsx(numeric_data, xlsx_path, selected_fields)
    
    print("\n분석 완료!")

def main():
    """메인 함수"""
    print("=" * 80)
    print("JSONL 수치 데이터 분석기")
    print("=" * 80)
    
    # 파일 경로 입력
    file_path = input("\nJSONL 파일 경로를 입력하세요: ").strip()
    
    # 특정 필드 선택 여부
    choice = input("특정 필드만 분석하시겠습니까? (y/n): ").strip().lower()
    
    selected_fields = None
    if choice == 'y':
        fields_input = input("분석할 필드명을 쉼표로 구분하여 입력하세요: ").strip()
        selected_fields = [f.strip() for f in fields_input.split(',') if f.strip()]
    
    # 저장 형식 선택
    save_choice = input("\n저장할 파일 형식을 선택하세요 (1: 둘 다, 2: TXT만, 3: XLSX만): ").strip()
    
    save_txt = save_choice in ['1', '2']
    save_xlsx = save_choice in ['1', '3']
    
    # 분석 실행
    try:
        analyze_jsonl(file_path, selected_fields, save_txt, save_xlsx)
    except FileNotFoundError:
        print(f"\n오류: 파일을 찾을 수 없습니다 - {file_path}")
    except json.JSONDecodeError as e:
        print(f"\n오류: JSON 파싱 실패 - {e}")
    except Exception as e:
        print(f"\n오류: {e}")

if __name__ == "__main__":
    main()
